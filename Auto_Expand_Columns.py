import openpyxl as op

# Creates a definition 
def auto_expand_columns(sheet):
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[op.utils.get_column_letter(column[0].column)].width = adjusted_width

# Load the workbook
workbook = op.load_workbook() #Input the location of the workbook that you would like to use

# Iterate through all sheets
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    auto_expand_columns(sheet)

# Save the modified workbook
workbook.save() # Input the location you would like to save to including the file name